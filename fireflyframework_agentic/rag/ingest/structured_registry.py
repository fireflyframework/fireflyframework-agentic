# Copyright 2026 Firefly Software Foundation
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"""Schema registry and schema discovery for structured ingestion.

``SchemaRegistry`` stores and retrieves ``TargetSchema`` objects in a
``_schemas`` table inside the corpus SQLite file.

``discover_schema`` uses ``create_extractor_agent`` (Claude) to infer a
``TargetSchema`` from headers and sample rows of a CSV or Excel file.
"""

from __future__ import annotations

import csv
import logging
from collections.abc import Awaitable, Callable
from pathlib import Path
from typing import Any

from fireflyframework_agentic.agents.base import FireflyAgent
from fireflyframework_agentic.rag.corpus import SqliteCorpus

from .structured_pipeline import (
    _HEADER_SCAN_ROWS,
    _normalize_sheet_name,
    _pick_header_row_idx,
)
from .structured_schema import SchemaFeedback, TableSpec, TargetSchema

# Re-exported for tests that import them from this module — moved to
# structured_pipeline.py so discovery and ingestion share one
# definition of "which row is the header in an Excel sheet."
__all__ = ["_HEADER_SCAN_ROWS", "_pick_header_row_idx"]

try:
    import openpyxl as _openpyxl

    _HAS_OPENPYXL = True
except ImportError:
    _openpyxl = None  # type: ignore[assignment]
    _HAS_OPENPYXL = False

log = logging.getLogger(__name__)

_SAMPLE_ROWS = 5

# File extensions the structured pipeline knows how to sample. Callers walking
# folders should filter to this set so non-tabular files (PPTX, PDF, DOCX, …)
# don't end up being read as CSV.
TABULAR_SUFFIXES: frozenset[str] = frozenset({".csv", ".xls", ".xlsx"})


def is_tabular_file(path: Path) -> bool:
    """Return True if *path*'s suffix is one the structured pipeline can sample."""
    return path.suffix.lower() in TABULAR_SUFFIXES


_SKILL = (
    "You are a data schema analyst. Given tabular data (headers + sample rows), "
    "infer the best relational schema.\n\n"
    "**Output contract — read this first:** ALWAYS return a TargetSchema with "
    "at least one TableSpec for every sheet / file in the input. Even when "
    "the data is messy (banner rows, merged cells, ambiguous column names, "
    "non-obvious types), produce your best structural interpretation — the "
    "user will review and correct it via the discover→review→ingest loop. "
    'An empty schema (``{}`` or ``{"tables": []}``) is never the right '
    "answer: it makes the tool fail with a validation error and the user "
    "cannot iterate from there. If a sheet looks truly empty after skipping "
    "banner rows, still emit a TableSpec for it with whatever columns you "
    "see and an inline comment in the column descriptions noting the "
    "ambiguity. The conservatism guidance below applies ONLY to the "
    "``unit`` field — every other field should reflect your best guess.\n\n"
    "**Naming:** Use snake_case for table and column names. "
    "For CSV files: table name = filename stem in snake_case. "
    "For Excel: each sheet becomes one table; table name = sheet name in snake_case.\n\n"
    "**Type inference:**\n"
    "- Integers only → integer\n"
    "- Numbers with decimals → float\n"
    "- true/false, yes/no, 0/1 → boolean\n"
    "- ISO dates, DD/MM/YYYY, MM-DD-YYYY → date\n"
    "- Datetimes with time component → datetime\n"
    "- Structured text (JSON-like) → json\n"
    "- Everything else → string\n\n"
    "**Unit inference (numeric columns):** When the column is integer or "
    "float AND there is a clear signal of what the values measure, set "
    "``unit`` to a short human-readable string. Leave ``unit`` null when "
    "there is no signal — guessing a unit is worse than admitting you "
    "don't know, because downstream answers will be confidently wrong.\n"
    "  Signals to use:\n"
    "  - Parenthesised hint in the header (case-insensitive):\n"
    '    ``Sales (USD)`` → unit="USD"; ``Revenue (USD millions)`` → '
    'unit="USD millions"; ``Achievement (%)`` → unit="percent"; '
    '``Duration (days)`` → unit="days".\n'
    "  - Currency suffix/prefix in the column name: ``amount_usd``, "
    '``price_eur``, ``revenue_gbp`` → unit="USD" / "EUR" / "GBP".\n'
    "  - Percent signals: header contains ``%``, ``pct``, ``percent``, "
    "``rate`` AND sample values look like percentages (0–100 or 0–1) "
    '→ unit="percent".\n'
    "  - Headcount-style integers: header is ``headcount``, ``fte``, "
    '``ftes``, ``employees``, ``staff`` → unit="headcount".\n'
    "  - Time/duration columns: header is ``days``, ``hours``, ``minutes``, "
    "``years`` → that word as the unit.\n"
    "  - Adjacent currency-code column (e.g. ``amount`` next to a ``currency`` "
    "column with one distinct value across the sample) → unit = that distinct "
    "value.\n"
    "  - Sheet-level / file-level header text in the sample mentioning units "
    "(e.g. a banner row ``Figures in EUR millions``) → apply to every "
    "money-shaped numeric column in that source.\n"
    "  Do NOT set ``unit`` on string / boolean / date / datetime / json "
    "columns, on primary keys, or on foreign keys — those aren't quantities. "
    "Do NOT invent a currency you cannot see; if the data is clearly a money "
    "amount but the currency is genuinely ambiguous, leave ``unit`` null. "
    "The answerer is instructed to flag that ambiguity rather than guess.\n\n"
    "**Nullability & primary keys:**\n"
    "- nullable: false only if every sample row has a non-empty value.\n"
    "- primary_key: true if the column looks like a unique identifier "
    "(named 'id' or ending in '_id', sequential integers with no duplicates).\n"
    "- Multiple columns can be marked primary_key — they form a composite "
    "primary key. Use this when the unique row identity is a combination of "
    "columns (e.g. (customer_id, period) on a monthly-metrics table, "
    "or (employee_id, route) on a per-route activity table where an "
    "employee can appear in multiple rows for different routes). When "
    "the sample shows duplicates on a single candidate ID column, the "
    "real key is almost always composite — pair it with the discriminating "
    "column rather than leaving the table keyless.\n"
    "- Leave every column primary_key=false when no combination of columns "
    "would yield a unique row. A table with no primary key is still a valid "
    "table and MUST appear in the schema; the absence of a PK is a property "
    "of the columns, not a reason to drop the table from the output.\n\n"
    "**Foreign keys (only when multiple tables are present):**\n"
    "- Set foreign_key to '<table>.<column>' when a column's values plausibly "
    "reference another table's primary key. Common signals: column ends in "
    "'_id' and a same-typed PK exists on another table whose name matches the "
    "prefix (e.g. customer_id → customers.id).\n"
    "- Do not invent references across unrelated tables; leave foreign_key "
    "null when the relationship is not clearly supported by names + types.\n"
    "- A column that is itself a primary_key (or part of a composite "
    "primary key) must NOT also be a foreign_key.\n\n"
    "**Multi-sheet Excel / multi-file folders:** One TableSpec per source "
    "(sheet or file). Skip sources where all sample rows are empty."
)

_DEFAULT_SCHEMA_MODEL = "anthropic:claude-sonnet-4-6"


class SchemaRegistry:
    """Stores and retrieves ``TargetSchema`` objects in the corpus SQLite file."""

    def __init__(self, corpus: SqliteCorpus) -> None:
        self._corpus = corpus

    async def initialise(self) -> None:
        await self._corpus.query(
            "CREATE TABLE IF NOT EXISTS _schemas (name TEXT PRIMARY KEY, schema_json TEXT NOT NULL)"
        )

    async def save(self, schema: TargetSchema) -> None:
        for table in schema.tables:
            single = TargetSchema(tables=[table])
            await self._corpus.query(
                "INSERT INTO _schemas (name, schema_json) VALUES (:name, :json) "
                "ON CONFLICT(name) DO UPDATE SET schema_json = excluded.schema_json",
                {"name": table.name, "json": single.model_dump_json()},
            )

    async def list_schemas(self) -> list[TargetSchema]:
        rows = await self._corpus.query("SELECT schema_json FROM _schemas")
        return [TargetSchema.model_validate_json(r["schema_json"]) for r in rows]


def _csv_sample(path: Path) -> str:
    try:
        with open(path, newline="", encoding="utf-8") as f:
            rows = list(csv.reader(f))[: _SAMPLE_ROWS + 1]
    except UnicodeDecodeError as exc:
        # Common case on Excel-exported CSVs from Windows: Latin-1 / CP1252.
        # We don't transparently fall back because that would mask real
        # encoding bugs; surface a clear hint instead.
        raise ValueError(
            f"could not decode {path.name} as UTF-8 (byte 0x{exc.object[exc.start]:02x} at offset {exc.start}). "
            "If the file was exported from Excel on Windows it may be Latin-1 / CP1252; "
            "re-save as UTF-8 or transcode (e.g. `iconv -f windows-1252 -t utf-8 in.csv > out.csv`, "
            "or in Python: `Path(p).write_bytes(Path(p).read_bytes().decode('cp1252').encode('utf-8'))`)."
        ) from exc
    name = path.stem.replace(" ", "_").replace("-", "_").lower()
    lines = [f"Table: {name}", f"Headers: {rows[0]}"]
    lines += [f"  {r}" for r in rows[1:]]
    return "\n".join(lines)


def _excel_sample(path: Path) -> str:
    if not _HAS_OPENPYXL:
        raise RuntimeError("openpyxl is required for Excel files: pip install openpyxl")
    wb = _openpyxl.load_workbook(path, read_only=True, data_only=True)  # type: ignore[union-attr]
    parts: list[str] = []
    # Pull enough rows for both header detection (first _HEADER_SCAN_ROWS)
    # and the sample window (_SAMPLE_ROWS + 1 after the picked header).
    scan_budget = _HEADER_SCAN_ROWS + _SAMPLE_ROWS + 1
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))[:scan_budget]
        if not all_rows:
            continue
        header_idx = _pick_header_row_idx(all_rows[:_HEADER_SCAN_ROWS])
        rows = all_rows[header_idx : header_idx + _SAMPLE_ROWS + 1]
        if not rows:
            continue
        # Skip sheets with no usable header — feeding the LLM
        # ``Headers: [None, None, None]`` blocks is what makes it return
        # ``{}`` on workbooks like the source workbook where one sheet
        # is essentially blank. ≥2 non-null cells AND ≥1 string is the
        # minimum bar for a header that's worth including.
        header = list(rows[0])
        non_null = [v for v in header if v is not None]
        string_count = sum(1 for v in non_null if isinstance(v, str))
        if len(non_null) < 2 or string_count == 0:
            log.warning(
                "schema discovery: skipping sheet %r in %s — no usable header row "
                "found within the first %d rows (header preview: %r)",
                sheet_name,
                path.name,
                _HEADER_SCAN_ROWS,
                header[:8],
            )
            continue
        name = _normalize_sheet_name(sheet_name)
        lines = [f"Sheet (table): {name}", f"Headers: {header}"]
        lines += [f"  {list(r)}" for r in rows[1:]]
        parts.append("\n".join(lines))
    return "\n\n".join(parts)


def _sample_for(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix in (".xls", ".xlsx"):
        return _excel_sample(path)
    if suffix == ".csv":
        return _csv_sample(path)
    raise ValueError(
        f"unsupported file type for structured ingest: {path.name!r} (suffix={suffix!r}). "
        f"Expected one of {sorted(TABULAR_SUFFIXES)}."
    )


def _multi_file_sample(paths: list[Path]) -> str:
    """Build a combined sample block for *paths* with one section per file."""
    parts: list[str] = []
    for p in paths:
        parts.append(f"File: {p.name}\n{_sample_for(p)}")
    return "\n\n".join(parts)


def _make_discovery_agent(model: str | Any) -> FireflyAgent[Any, TargetSchema]:
    """Build the schema-discovery agent.

    Uses ``FireflyAgent`` directly with ``_SKILL`` as the *full*
    instructions rather than going through ``create_extractor_agent``.
    The extractor template's preamble ("If a field cannot be found,
    return null. Do not infer or hallucinate values.") conflicts with
    schema discovery, where ``tables`` cannot be null and the model
    must produce *some* best-effort interpretation of every sheet —
    that conflict was the proximate cause of empty-``{}`` outputs on
    messy real-world workbooks.
    """
    return FireflyAgent(
        "schema_discovery",
        model=model,
        instructions=_SKILL,
        output_type=TargetSchema,
        description="Infers a relational TargetSchema from tabular samples.",
        tags=("rag", "ingest", "structured", "schema"),
        auto_register=False,
    )


_RETRY_NUDGE = (
    "\n\nIMPORTANT: your previous response contained no tables. The "
    "output contract above is binding — produce a non-empty TargetSchema "
    "with one TableSpec per sheet/file in the sample, even on messy or "
    "ambiguous data. Empty is not an option."
)


async def _run_with_retry_on_empty(agent: Any, prompt: str) -> TargetSchema:
    """Run the discovery agent; on an empty result, retry once with a
    short nudge. The LLM is non-deterministic on borderline-messy inputs
    (the same sample produces a non-empty schema on most runs and ``{}``
    on a minority). One retry is cheap and turns a flaky pass rate into
    a reliable one without changing the failure semantics of the safety
    net: if the second attempt is also empty, we still raise the
    diagnostic ValueError so the user sees a clear actionable error
    rather than silent corruption.
    """
    result = await agent.run(prompt)
    schema: TargetSchema = result.output
    if schema.tables:
        return schema
    log.warning("schema discovery returned empty schema; retrying once with explicit nudge")
    result = await agent.run(prompt + _RETRY_NUDGE)
    return result.output  # caller checks .tables and raises if still empty


def _empty_schema_diagnostic(
    label: str,
    sample: str,
    *,
    previous_schema: TargetSchema | None = None,
    corrections: str = "",
) -> str:
    """Build a human-friendly error for ``schema.tables == []`` cases.

    Includes a truncated sample so the user can see what the model saw
    without flooding the error stream (real samples are ~10 KB). When
    the empty schema came out of a *refinement* call (previous_schema
    non-empty), the diagnostic frames it differently: the model
    discarded the validated baseline rather than failing to find one.
    """
    preview = sample[:1800] + ("…" if len(sample) > 1800 else "")
    if previous_schema is not None and previous_schema.tables:
        prior_names = ", ".join(t.name for t in previous_schema.tables)
        return (
            f"schema discovery refinement returned an empty schema for "
            f"{label}, discarding the {len(previous_schema.tables)}-table "
            f"baseline you already validated ({prior_names}). The model "
            "should have carried those tables forward and only edited the "
            "ones your corrections targeted. "
            "Fix options: (1) the MCP host can fall back to the previous "
            "schema and apply the corrections programmatically (the "
            "TargetSchema dict is editable); (2) re-issue "
            "discover_corpus_schema with a narrower correction string "
            "that targets one table at a time. "
            f"\n\nCorrections that triggered this:\n{corrections}"
            f"\n\nSample sent to the model (truncated to 1800 chars):\n{preview}"
        )
    return (
        f"schema discovery produced an empty schema for {label}. "
        "The model could not infer any tables from the sampled rows — "
        "this usually means the sheets are non-tabular (pivot reports, "
        "dashboards, merged-cell layouts) or the real headers sit past "
        "the scan window. "
        "Fix options: (1) re-save each sheet as a flat CSV with headers "
        "in row 1 and retry; (2) call discover_corpus_schema again with "
        "`corrections=<free-text description of the table structure>` so "
        "the model gets explicit guidance.\n\n"
        f"Sample sent to the model (truncated to 1800 chars):\n{preview}"
    )


async def discover_schema(
    path: Path,
    *,
    model: str = _DEFAULT_SCHEMA_MODEL,
    corrections: str = "",
    previous_schema: TargetSchema | None = None,
) -> TargetSchema:
    """Infer a ``TargetSchema`` from *path* (CSV or Excel).

    When *corrections* and *previous_schema* are provided the agent is asked to
    refine *previous_schema* according to the supplied corrections.

    Raises ``ValueError`` (wrapped to ``ToolError`` upstream) when the
    LLM returns an empty schema — pydantic-ai no longer retries this
    case because ``TargetSchema.tables`` defaults to ``[]``.
    """
    agent = _make_discovery_agent(model)
    sample = _sample_for(path)
    if corrections and previous_schema is not None:
        prompt = (
            f"File: {path.name}\n\n{sample}\n\n"
            f"Previous schema attempt:\n{previous_schema.model_dump_json(indent=2)}\n\n"
            f"User corrections:\n{corrections}\n\n"
            "Produce a corrected schema addressing the user's corrections. "
            "Refinement contract: your output MUST contain every table from "
            "the previous schema that the corrections do not explicitly ask "
            "you to drop. If a correction says 'drop X' or 'remove X', omit "
            "that table; otherwise carry it forward — modified per the "
            "corrections if they target it, unchanged otherwise. Returning "
            "an empty schema during refinement is never correct: the user "
            "already validated the previous schema, and the corrections are "
            "edits on top of it, not a signal to start over."
        )
    else:
        prompt = f"File: {path.name}\n\n{sample}"
    schema = await _run_with_retry_on_empty(agent, prompt)
    if not schema.tables:
        raise ValueError(
            _empty_schema_diagnostic(
                path.name,
                sample,
                previous_schema=previous_schema,
                corrections=corrections,
            )
        )
    return schema


async def discover_schema_for_paths(
    paths: list[Path],
    *,
    model: str = _DEFAULT_SCHEMA_MODEL,
    corrections: str = "",
    previous_schema: TargetSchema | None = None,
) -> TargetSchema:
    """Infer a ``TargetSchema`` across *paths* in a single LLM call.

    The combined sample lets the agent propose cross-file foreign keys
    (e.g. ``billing_ledger.customer_id`` → ``customers.id``) that per-file
    discovery cannot see. Returns one ``TargetSchema`` containing one
    ``TableSpec`` per source (CSV file, or sheet for Excel inputs).

    When *corrections* and *previous_schema* are provided the agent is asked
    to refine *previous_schema* according to the supplied free-text feedback.

    Raises ``ValueError`` when the LLM returns an empty schema. See
    :func:`discover_schema` for the rationale.
    """
    if not paths:
        return TargetSchema(tables=[])
    if len(paths) == 1:
        return await discover_schema(paths[0], model=model, corrections=corrections, previous_schema=previous_schema)
    # Multi-file: prefer per-file discovery and merge. A single combined
    # LLM call (one prompt, every sheet of every file concatenated) was
    # the original design — it let the model propose cross-file foreign
    # keys — but it is fragile on real workbooks: one messy sheet (dual
    # pivot, stacked sub-tables) anywhere in the combined sample makes
    # the model return ``{}`` for the *entire* multi-file output. Per-file
    # discovery isolates that failure mode to the offending file, and
    # cross-file FK relationships can be added by the user via a
    # refinement round with explicit corrections.
    if corrections and previous_schema is not None:
        # Refinement is intrinsically multi-file (it operates on the
        # union schema), so keep the original combined-prompt path here.
        agent = _make_discovery_agent(model)
        sample = _multi_file_sample(paths)
        prompt = (
            f"You are given {len(paths)} tabular sources.\n\n{sample}\n\n"
            f"Previous schema attempt:\n{previous_schema.model_dump_json(indent=2)}\n\n"
            f"User corrections:\n{corrections}\n\n"
            "Produce a corrected schema addressing the user's corrections. "
            "Re-evaluate foreign_key relationships in light of the feedback. "
            "Refinement contract: your output MUST contain every table from "
            "the previous schema that the corrections do not explicitly ask "
            "you to drop. If a correction says 'drop X' or 'remove X', omit "
            "that table; otherwise carry it forward — modified per the "
            "corrections if they target it, unchanged otherwise. Returning "
            "an empty schema during refinement is never correct."
        )
        schema = await _run_with_retry_on_empty(agent, prompt)
        if not schema.tables:
            label = (
                f"{len(paths)} tabular sources ({', '.join(p.name for p in paths[:5])}{'…' if len(paths) > 5 else ''})"
            )
            raise ValueError(
                _empty_schema_diagnostic(
                    label,
                    sample,
                    previous_schema=previous_schema,
                    corrections=corrections,
                )
            )
        return schema
    # Initial discovery: per-file + merge.
    merged: list[TableSpec] = []
    for p in paths:
        s = await discover_schema(p, model=model)
        merged.extend(s.tables)
    return TargetSchema(tables=merged)


async def discover_schema_interactive(
    path: Path,
    *,
    on_review: Callable[[TargetSchema], Awaitable[SchemaFeedback]],
    model: str = _DEFAULT_SCHEMA_MODEL,
    max_rounds: int = 5,
) -> TargetSchema:
    """Infer a TargetSchema from *path* with iterative user refinement.

    Each round: infer schema → call *on_review* with the result.
    If feedback.approved is True, return immediately.
    If False, inject feedback.corrections into the next inference prompt.
    After *max_rounds*, return the last schema regardless of approval.
    """
    corrections: str = ""
    schema: TargetSchema | None = None

    for _ in range(max_rounds):
        schema = await discover_schema(
            path,
            model=model,
            corrections=corrections,
            previous_schema=schema,
        )
        feedback = await on_review(schema)
        if feedback.approved:
            return schema
        corrections = feedback.corrections

    assert schema is not None
    # Operator rejected every round but we still have a schema in hand;
    # surface this as a warning rather than silently proceeding under a
    # schema the user explicitly disapproved of.
    log.warning(
        "schema discovery for %s exhausted %d rounds without operator approval; "
        "ingesting under the last inferred schema (last corrections=%r)",
        path,
        max_rounds,
        corrections,
    )
    return schema
