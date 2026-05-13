# Copyright 2026 Firefly Software Foundation
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"""Structured data ingestion pipeline.

Reads rows from a CSV or Excel file and inserts them into SQLite tables
according to a ``TargetSchema``.  No LLM calls — pure Python.
"""

from __future__ import annotations

import asyncio
import csv
import logging
import re
import sqlite3
from pathlib import Path
from typing import TYPE_CHECKING, Any, Literal, TypedDict

from .structured_schema import ColumnType, TableSpec, TargetSchema

if TYPE_CHECKING:
    from fireflyframework_agentic.storage import DatabaseStore

log = logging.getLogger(__name__)


class TableIngestResult(TypedDict):
    """Per-table outcome returned by :func:`ingest_structured`.

    ``status``:
      - ``"success"`` — every row inserted, no errors.
      - ``"partial"`` — at least one row failed (UNIQUE/NOT NULL/type
        violation) but the rest were committed. This is the case we
        hit on real workbooks where a handful of placeholder rows
        (e.g. ``prid = '-'`` vacancy markers) conflict with a composite
        primary key — atomic rollback used to lose ~680 valid rows to
        protect 13 conflicting ones, which is the wrong trade-off in
        practice.
      - ``"failed"`` — zero rows inserted (every row errored, or the
        sheet was empty after filtering).

    ``errors`` carries one entry per failing row in
    ``f"row {row_num}: {sqlite_message}"`` form, regardless of status.
    """

    status: Literal["success", "partial", "failed"]
    inserted: int
    errors: list[str]


try:
    import openpyxl as _openpyxl

    _HAS_OPENPYXL = True
except ImportError:
    _openpyxl = None  # type: ignore[assignment]
    _HAS_OPENPYXL = False

_SQL_TYPES: dict[ColumnType, str] = {
    ColumnType.string: "TEXT",
    ColumnType.integer: "INTEGER",
    ColumnType.float_: "REAL",
    ColumnType.boolean: "INTEGER",
    ColumnType.date: "TEXT",
    ColumnType.datetime: "TEXT",
    ColumnType.json: "TEXT",
}


def _normalize_sheet_name(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", name.lower().replace("&", "and")).strip("_")


def _normalize_col(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(name).lower().replace("&", "and")).strip("_")


# How far down a sheet to look for the real header row. Real-world Excel
# files frequently bury headers under ~10 single-cell title / legend / spacer
# rows (e.g. an archive sheet in the the test workbook puts the header at
# row 8). Both discovery (_excel_sample) and ingestion (_read_rows) must
# use the same value or they drift — discovery sees the right schema while
# ingestion silently drops every row.
_HEADER_SCAN_ROWS = 20


def _pick_header_row_idx(candidate_rows: list[tuple[Any, ...]]) -> int:
    """Choose the row in *candidate_rows* most likely to be the real header.

    Single source of truth for header detection — used by ``_excel_sample``
    during schema discovery AND by ``_read_rows`` during ingestion. Keeping
    these two code paths in sync is load-bearing: when they pick different
    rows, discovery infers a correct schema but ingestion can't find the
    columns to populate it, and every data row gets silently filtered out
    by the NOT NULL pre-filter in ``_load_rows``.

    Heuristic, in order:
      1. Skip empty / one-cell title rows (a real header has ≥2 cells).
      2. Among multi-cell *string-dominated* rows (>50% of non-null cells
         are strings — real Excel headers are almost always strings, not
         integers), pick the one with the most non-null cells. Real
         headers are wide; decorative two-cell titles like
         ``['DECORATIVE TITLE', '(en blanco)']`` lose to the
         5-column real header at row 2.
      3. Fall back to the first multi-cell row.

    Returns 0 when no row passes any test (so the existing behaviour for
    well-formed sheets is preserved).

    The caller is expected to pass at most ``_HEADER_SCAN_ROWS`` rows;
    that bound governs how far the picker will look.
    """
    multi_cell = [(i, r) for i, r in enumerate(candidate_rows) if sum(1 for v in r if v is not None) >= 2]
    if not multi_cell:
        return 0
    string_dominated: list[tuple[int, int]] = []  # (row_idx, non_null_count)
    for i, r in multi_cell:
        non_null = [v for v in r if v is not None]
        string_count = sum(1 for v in non_null if isinstance(v, str))
        if string_count * 2 > len(non_null):
            string_dominated.append((i, len(non_null)))
    if string_dominated:
        # Widest row wins, earliest row breaks ties.
        return min(string_dominated, key=lambda t: (-t[1], t[0]))[0]
    return multi_cell[0][0]


def _read_rows(path: Path, table_name: str) -> tuple[list[str], list[list[Any]]]:
    suffix = path.suffix.lower()
    if suffix in (".xls", ".xlsx"):
        if not _HAS_OPENPYXL:
            raise RuntimeError("openpyxl is required for Excel files: pip install openpyxl")
        wb = _openpyxl.load_workbook(path, read_only=True, data_only=True)  # type: ignore[union-attr]
        sheet_name = next(
            (s for s in wb.sheetnames if _normalize_sheet_name(s) == table_name),
            None,
        )
        if sheet_name is None:
            raise KeyError(f"No sheet matching table {table_name!r} in {path.name} (sheets: {wb.sheetnames})")
        all_rows = list(wb[sheet_name].iter_rows(values_only=True))
        header_idx = _pick_header_row_idx(all_rows[:_HEADER_SCAN_ROWS])
        return [str(h) if h is not None else "" for h in all_rows[header_idx]], [
            list(r) for r in all_rows[header_idx + 1 :]
        ]
    try:
        with open(path, newline="", encoding="utf-8") as f:
            rows = list(csv.reader(f))
    except UnicodeDecodeError as exc:
        raise ValueError(
            f"could not decode {path.name} as UTF-8 (byte 0x{exc.object[exc.start]:02x} "
            f"at offset {exc.start}). If the file was exported from Excel on Windows "
            "it may be Latin-1 / CP1252; re-save as UTF-8 or transcode "
            "(e.g. `iconv -f windows-1252 -t utf-8 in.csv > out.csv`, or in Python: "
            "`Path(p).write_bytes(Path(p).read_bytes().decode('cp1252').encode('utf-8'))`)."
        ) from exc
    return rows[0], rows[1:]


def _load_rows(path: Path, schema: TargetSchema) -> dict[str, list[dict[str, Any]] | None]:
    """Return table_name → row dicts, or None when required columns are absent.

    Matching strategy (in order):
    1. Normalized name match — ``FY2022`` matches schema column ``fy2022``.
    2. Positional fallback — when the LLM semantically renamed a column
       (e.g. "Fiscal Year" → ``line_item``), fall back to aligning schema
       columns to file columns by position, provided the counts match.
    """
    rows_by_table: dict[str, list[dict[str, Any]] | None] = {}
    for table in schema.tables:
        headers, raw_rows = _read_rows(path, table.name)
        # Build normalized-name → index map (skip empty/None header cells).
        norm_header_idx: dict[str, int] = {}
        for i, h in enumerate(headers):
            nk = _normalize_col(h)
            if nk and nk not in norm_header_idx:
                norm_header_idx[nk] = i

        col_names = [c.name for c in table.columns]
        missing = [c for c in col_names if _normalize_col(c) not in norm_header_idx]

        if missing:
            # Positional fallback: align schema column i to file column i.
            # First try matching the total column count (including None-header cells,
            # common in multi-section sheets like dashboards).  If that doesn't fit,
            # try matching only non-empty header positions.
            if len(headers) >= len(col_names):
                col_idx = {c: i for i, c in enumerate(col_names)}
            else:
                non_empty_headers = [h for h in headers if h and h.strip()]
                if len(non_empty_headers) >= len(col_names):
                    col_idx = {c: headers.index(non_empty_headers[i]) for i, c in enumerate(col_names)}
                else:
                    rows_by_table[table.name] = None
                    continue
        else:
            col_idx = {c: norm_header_idx[_normalize_col(c)] for c in col_names}

        # Pre-filter section-header / spacer rows (e.g. ["REVENUE", None, …])
        # that would violate NOT NULL constraints. Capture loop-local copies of
        # col_idx and not_null_set to avoid B023 late-binding issues.
        not_null_set = frozenset(c.name for c in table.columns if not c.nullable or c.primary_key)
        _ci = col_idx
        _nn = not_null_set
        rows_by_table[table.name] = [
            {c: (row[_ci[c]] if _ci[c] < len(row) else None) for c in col_names}
            for row in raw_rows
            if all(_ci[c] >= len(row) or row[_ci[c]] is not None for c in _nn)
        ]
    return rows_by_table


def _quote(name: str) -> str:
    return '"' + name.replace('"', '""') + '"'


def _sync_ingest_table(
    db_path: Path,
    table_spec: TableSpec,
    rows: list[dict[str, Any]],
) -> TableIngestResult:
    # isolation_level=None puts the driver in autocommit so the BEGIN /
    # COMMIT / ROLLBACK statements below are real SQL transactions rather
    # than driver-level no-ops. busy_timeout=30000 PRAGMA defends against
    # the rare case where another sqlite3 sidecar overrides the
    # constructor's `timeout=` arg on this same connection.
    conn = sqlite3.connect(db_path, timeout=30.0, isolation_level=None)
    conn.execute("PRAGMA busy_timeout = 30000")
    try:
        pk_cols = [c for c in table_spec.columns if c.primary_key]
        # Single-col PK keeps the inline ``PRIMARY KEY`` form so an
        # ``INTEGER PRIMARY KEY`` column remains the SQLite rowid alias
        # (the table-level form below would forfeit that). With ≥2 PK
        # columns the row identity is composite and we emit a
        # ``PRIMARY KEY (col_a, col_b, …)`` clause at the end of the
        # column list. This is what unblocks tables like
        # ``monthly_table`` where ``prid`` repeats across rows but
        # ``(prid, ruta_num)`` is genuinely unique.
        composite_pk = len(pk_cols) > 1
        col_defs: list[str] = []
        fk_defs: list[str] = []
        for col in table_spec.columns:
            parts = [_quote(col.name), _SQL_TYPES[col.type]]
            if col.primary_key and not composite_pk:
                parts.append("PRIMARY KEY")
            elif not col.nullable:
                parts.append("NOT NULL")
            col_defs.append(" ".join(parts))
            if col.foreign_key:
                ref_table, _, ref_column = col.foreign_key.partition(".")
                if ref_table and ref_column:
                    fk_defs.append(
                        f"FOREIGN KEY ({_quote(col.name)}) REFERENCES {_quote(ref_table)} ({_quote(ref_column)})"
                    )
        pk_def = [f"PRIMARY KEY ({', '.join(_quote(c.name) for c in pk_cols)})"] if composite_pk else []
        all_defs = col_defs + pk_def + fk_defs
        conn.execute(f"CREATE TABLE IF NOT EXISTS {_quote(table_spec.name)} ({', '.join(all_defs)})")
        col_names = [c.name for c in table_spec.columns]
        placeholders = ", ".join("?" for _ in col_names)
        col_names_str = ", ".join(_quote(c) for c in col_names)
        errors: list[str] = []
        inserted = 0
        conn.execute("BEGIN")
        try:
            for row_num, row in enumerate(rows, start=2):
                values = [row.get(c) for c in col_names]
                # Skip section-header rows: all values are None (common in Excel).
                if all(v is None for v in values):
                    continue
                try:
                    conn.execute(
                        f"INSERT INTO {_quote(table_spec.name)} ({col_names_str}) VALUES ({placeholders})",
                        values,
                    )
                    inserted += 1
                except sqlite3.Error as exc:
                    errors.append(f"row {row_num}: {exc}")
            # Commit whatever inserted successfully; per-row errors are
            # reported but don't roll back valid rows. ``failed`` is
            # reserved for "every row that was attempted errored" — an
            # empty CSV (header only, no data rows) is still a success
            # with inserted=0 because there was nothing to fail on.
            conn.execute("COMMIT")
            if errors and inserted == 0:
                return {"status": "failed", "inserted": 0, "errors": errors}
            if errors:
                return {"status": "partial", "inserted": inserted, "errors": errors}
            return {"status": "success", "inserted": inserted, "errors": []}
        except BaseException:
            try:
                conn.execute("ROLLBACK")
            except sqlite3.Error as rb_exc:
                log.warning(
                    "rollback failed during error recovery on table %s: %s",
                    table_spec.name,
                    rb_exc,
                )
            raise
    finally:
        conn.close()


def _order_tables_by_fk(tables: list[TableSpec]) -> list[TableSpec]:
    """Return *tables* topologically sorted so referenced tables are created first.

    Cycles or references to unknown tables are ignored — the offending table
    just keeps its original position. SQLite will surface a real FK error at
    insert time if the reference is wrong.
    """
    by_name = {t.name: t for t in tables}
    ordered: list[TableSpec] = []
    seen: set[str] = set()

    def visit(t: TableSpec, on_path: set[str]) -> None:
        if t.name in seen or t.name in on_path:
            return
        on_path.add(t.name)
        for col in t.columns:
            if col.foreign_key:
                ref_table = col.foreign_key.split(".", 1)[0]
                if ref_table in by_name and ref_table != t.name:
                    visit(by_name[ref_table], on_path)
        on_path.discard(t.name)
        seen.add(t.name)
        ordered.append(t)

    for t in tables:
        visit(t, set())
    return ordered


async def ingest_structured(
    path: Path,
    db_store: DatabaseStore,
    schema: TargetSchema,
) -> dict[str, TableIngestResult]:
    """Insert rows from *path* into the SQLite file owned by *db_store*.

    Routes through ``db_store.for_write()`` so the asyncio + sentinel locks
    inside the storage backend serialise this writer with anything else
    talking to the same file. Returns ``{table_name: TableIngestResult}``
    (see :class:`TableIngestResult` for the per-table fields).
    """
    rows_by_table = _load_rows(path, schema)
    results: dict[str, TableIngestResult] = {}
    async with db_store.for_write() as session:
        loop = asyncio.get_running_loop()
        for table_spec in _order_tables_by_fk(schema.tables):
            rows = rows_by_table.get(table_spec.name)
            if rows is None:
                results[table_spec.name] = {
                    "status": "failed",
                    "inserted": 0,
                    "errors": [f"missing columns for table {table_spec.name!r}"],
                }
                continue
            results[table_spec.name] = await loop.run_in_executor(
                None, _sync_ingest_table, session.path, table_spec, rows
            )
    return results
