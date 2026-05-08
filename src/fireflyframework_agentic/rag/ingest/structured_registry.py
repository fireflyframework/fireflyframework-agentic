# Copyright 2026 Firefly Software Foundation
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"""Schema registry and schema discovery for structured ingestion.

``SchemaRegistry`` stores and retrieves ``TargetSchema`` objects in a
``_schemas`` table inside the corpus SQLite file.

``discover_schema`` uses ``create_extractor_agent`` (Claude) to infer a
``TargetSchema`` from headers and sample rows of a CSV or Excel file.
"""

from __future__ import annotations

import csv
import logging
from collections.abc import Awaitable, Callable
from pathlib import Path

from fireflyframework_agentic.agents.templates import create_extractor_agent
from fireflyframework_agentic.rag.corpus import SqliteCorpus

from .structured_pipeline import _normalize_sheet_name
from .structured_schema import SchemaFeedback, TargetSchema

try:
    import openpyxl as _openpyxl

    _HAS_OPENPYXL = True
except ImportError:
    _openpyxl = None  # type: ignore[assignment]
    _HAS_OPENPYXL = False

log = logging.getLogger(__name__)

_SAMPLE_ROWS = 5

# File extensions the structured pipeline knows how to sample. Callers walking
# folders should filter to this set so non-tabular files (PPTX, PDF, DOCX, …)
# don't end up being read as CSV.
TABULAR_SUFFIXES: frozenset[str] = frozenset({".csv", ".xls", ".xlsx"})


def is_tabular_file(path: Path) -> bool:
    """Return True if *path*'s suffix is one the structured pipeline can sample."""
    return path.suffix.lower() in TABULAR_SUFFIXES


_SKILL = (
    "You are a data schema analyst. Given tabular data (headers + sample rows), "
    "infer the best relational schema.\n\n"
    "**Naming:** Use snake_case for table and column names. "
    "For CSV files: table name = filename stem in snake_case. "
    "For Excel: each sheet becomes one table; table name = sheet name in snake_case.\n\n"
    "**Type inference:**\n"
    "- Integers only → integer\n"
    "- Numbers with decimals → float\n"
    "- true/false, yes/no, 0/1 → boolean\n"
    "- ISO dates, DD/MM/YYYY, MM-DD-YYYY → date\n"
    "- Datetimes with time component → datetime\n"
    "- Structured text (JSON-like) → json\n"
    "- Everything else → string\n\n"
    "**Nullability & primary keys:**\n"
    "- nullable: false only if every sample row has a non-empty value.\n"
    "- primary_key: true if the column looks like a unique identifier "
    "(named 'id' or ending in '_id', sequential integers with no duplicates).\n"
    "- At most one primary key per table.\n\n"
    "**Foreign keys (only when multiple tables are present):**\n"
    "- Set foreign_key to '<table>.<column>' when a column's values plausibly "
    "reference another table's primary key. Common signals: column ends in "
    "'_id' and a same-typed PK exists on another table whose name matches the "
    "prefix (e.g. customer_id → customers.id).\n"
    "- Do not invent references across unrelated tables; leave foreign_key "
    "null when the relationship is not clearly supported by names + types.\n"
    "- A column that is itself a primary_key must NOT also be a foreign_key.\n\n"
    "**Multi-sheet Excel / multi-file folders:** One TableSpec per source "
    "(sheet or file). Skip sources where all sample rows are empty."
)

_DEFAULT_SCHEMA_MODEL = "anthropic:claude-sonnet-4-6"


class SchemaRegistry:
    """Stores and retrieves ``TargetSchema`` objects in the corpus SQLite file."""

    def __init__(self, corpus: SqliteCorpus) -> None:
        self._corpus = corpus

    async def initialise(self) -> None:
        await self._corpus.query(
            "CREATE TABLE IF NOT EXISTS _schemas (name TEXT PRIMARY KEY, schema_json TEXT NOT NULL)"
        )

    async def save(self, schema: TargetSchema) -> None:
        for table in schema.tables:
            single = TargetSchema(tables=[table])
            await self._corpus.query(
                "INSERT INTO _schemas (name, schema_json) VALUES (:name, :json) "
                "ON CONFLICT(name) DO UPDATE SET schema_json = excluded.schema_json",
                {"name": table.name, "json": single.model_dump_json()},
            )

    async def list_schemas(self) -> list[TargetSchema]:
        rows = await self._corpus.query("SELECT schema_json FROM _schemas")
        return [TargetSchema.model_validate_json(r["schema_json"]) for r in rows]


def _csv_sample(path: Path) -> str:
    try:
        with open(path, newline="", encoding="utf-8") as f:
            rows = list(csv.reader(f))[: _SAMPLE_ROWS + 1]
    except UnicodeDecodeError as exc:
        # Common case on Excel-exported CSVs from Windows: Latin-1 / CP1252.
        # We don't transparently fall back because that would mask real
        # encoding bugs; surface a clear hint instead.
        raise ValueError(
            f"could not decode {path.name} as UTF-8 (byte 0x{exc.object[exc.start]:02x} at offset {exc.start}). "
            "If the file was exported from Excel on Windows it may be Latin-1 / CP1252; "
            "re-save as UTF-8 or run `iconv -f windows-1252 -t utf-8` on it."
        ) from exc
    name = path.stem.replace(" ", "_").replace("-", "_").lower()
    lines = [f"Table: {name}", f"Headers: {rows[0]}"]
    lines += [f"  {r}" for r in rows[1:]]
    return "\n".join(lines)


def _excel_sample(path: Path) -> str:
    if not _HAS_OPENPYXL:
        raise RuntimeError("openpyxl is required for Excel files: pip install openpyxl")
    wb = _openpyxl.load_workbook(path, read_only=True, data_only=True)  # type: ignore[union-attr]
    parts: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))[: _SAMPLE_ROWS + 2]
        if not all_rows:
            continue
        # Skip title rows (rows with only one non-null cell) to show real headers.
        header_idx = next(
            (i for i, r in enumerate(all_rows[:5]) if sum(1 for v in r if v is not None) >= 2),
            0,
        )
        rows = all_rows[header_idx:]
        if not rows:
            continue
        name = _normalize_sheet_name(sheet_name)
        lines = [f"Sheet (table): {name}", f"Headers: {list(rows[0])}"]
        lines += [f"  {list(r)}" for r in rows[1:]]
        parts.append("\n".join(lines))
    return "\n\n".join(parts)


def _sample_for(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix in (".xls", ".xlsx"):
        return _excel_sample(path)
    if suffix == ".csv":
        return _csv_sample(path)
    raise ValueError(
        f"unsupported file type for structured ingest: {path.name!r} (suffix={suffix!r}). "
        f"Expected one of {sorted(TABULAR_SUFFIXES)}."
    )


def _multi_file_sample(paths: list[Path]) -> str:
    """Build a combined sample block for *paths* with one section per file."""
    parts: list[str] = []
    for p in paths:
        parts.append(f"File: {p.name}\n{_sample_for(p)}")
    return "\n\n".join(parts)


async def discover_schema(
    path: Path,
    *,
    model: str = _DEFAULT_SCHEMA_MODEL,
    corrections: str = "",
    previous_schema: TargetSchema | None = None,
) -> TargetSchema:
    """Infer a ``TargetSchema`` from *path* (CSV or Excel).

    When *corrections* and *previous_schema* are provided the agent is asked to
    refine *previous_schema* according to the supplied corrections.
    """
    agent = create_extractor_agent(
        TargetSchema,
        name="schema_discovery",
        model=model,
        extra_instructions=_SKILL,
        auto_register=False,
    )
    sample = _sample_for(path)
    if corrections and previous_schema is not None:
        prompt = (
            f"File: {path.name}\n\n{sample}\n\n"
            f"Previous schema attempt:\n{previous_schema.model_dump_json(indent=2)}\n\n"
            f"User corrections:\n{corrections}\n\n"
            "Produce a corrected schema addressing the user's corrections."
        )
    else:
        prompt = f"File: {path.name}\n\n{sample}"
    result = await agent.run(prompt)
    return result.output


async def discover_schema_for_paths(
    paths: list[Path],
    *,
    model: str = _DEFAULT_SCHEMA_MODEL,
    corrections: str = "",
    previous_schema: TargetSchema | None = None,
) -> TargetSchema:
    """Infer a ``TargetSchema`` across *paths* in a single LLM call.

    The combined sample lets the agent propose cross-file foreign keys
    (e.g. ``billing_ledger.customer_id`` → ``customers.id``) that per-file
    discovery cannot see. Returns one ``TargetSchema`` containing one
    ``TableSpec`` per source (CSV file, or sheet for Excel inputs).

    When *corrections* and *previous_schema* are provided the agent is asked
    to refine *previous_schema* according to the supplied free-text feedback.
    """
    if not paths:
        return TargetSchema(tables=[])
    if len(paths) == 1:
        return await discover_schema(paths[0], model=model, corrections=corrections, previous_schema=previous_schema)
    agent = create_extractor_agent(
        TargetSchema,
        name="schema_discovery",
        model=model,
        extra_instructions=_SKILL,
        auto_register=False,
    )
    sample = _multi_file_sample(paths)
    if corrections and previous_schema is not None:
        prompt = (
            f"You are given {len(paths)} tabular sources.\n\n{sample}\n\n"
            f"Previous schema attempt:\n{previous_schema.model_dump_json(indent=2)}\n\n"
            f"User corrections:\n{corrections}\n\n"
            "Produce a corrected schema addressing the user's corrections. "
            "Re-evaluate foreign_key relationships in light of the feedback."
        )
    else:
        prompt = (
            f"You are given {len(paths)} tabular sources. Infer one TableSpec "
            f"per source and propose foreign_key relationships where the data "
            f"clearly supports them.\n\n{sample}"
        )
    result = await agent.run(prompt)
    return result.output


async def discover_schema_interactive(
    path: Path,
    *,
    on_review: Callable[[TargetSchema], Awaitable[SchemaFeedback]],
    model: str = _DEFAULT_SCHEMA_MODEL,
    max_rounds: int = 5,
) -> TargetSchema:
    """Infer a TargetSchema from *path* with iterative user refinement.

    Each round: infer schema → call *on_review* with the result.
    If feedback.approved is True, return immediately.
    If False, inject feedback.corrections into the next inference prompt.
    After *max_rounds*, return the last schema regardless of approval.
    """
    corrections: str = ""
    schema: TargetSchema | None = None

    for _ in range(max_rounds):
        schema = await discover_schema(
            path,
            model=model,
            corrections=corrections,
            previous_schema=schema,
        )
        feedback = await on_review(schema)
        if feedback.approved:
            return schema
        corrections = feedback.corrections

    assert schema is not None
    # Operator rejected every round but we still have a schema in hand;
    # surface this as a warning rather than silently proceeding under a
    # schema the user explicitly disapproved of.
    log.warning(
        "schema discovery for %s exhausted %d rounds without operator approval; "
        "ingesting under the last inferred schema (last corrections=%r)",
        path,
        max_rounds,
        corrections,
    )
    return schema
