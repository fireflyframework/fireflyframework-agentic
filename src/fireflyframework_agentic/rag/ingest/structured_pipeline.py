# Copyright 2026 Firefly Software Foundation
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"""Structured data ingestion pipeline.

Reads rows from a CSV or Excel file and inserts them into SQLite tables
according to a ``TargetSchema``.  No LLM calls — pure Python.
"""

from __future__ import annotations

import asyncio
import csv
import re
import sqlite3
from pathlib import Path
from typing import Any

from .structured_schema import ColumnType, TableSpec, TargetSchema

try:
    import openpyxl as _openpyxl

    _HAS_OPENPYXL = True
except ImportError:
    _openpyxl = None  # type: ignore[assignment]
    _HAS_OPENPYXL = False

_SQL_TYPES: dict[ColumnType, str] = {
    ColumnType.string: "TEXT",
    ColumnType.integer: "INTEGER",
    ColumnType.float_: "REAL",
    ColumnType.boolean: "INTEGER",
    ColumnType.date: "TEXT",
    ColumnType.datetime: "TEXT",
    ColumnType.json: "TEXT",
}


def _normalize_sheet_name(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", name.lower().replace("&", "and")).strip("_")


def _normalize_col(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(name).lower().replace("&", "and")).strip("_")


def _find_header_row(rows: list[tuple[Any, ...]]) -> int:
    """Return the index of the first row that looks like real headers.

    A title row typically has exactly one non-null cell; the actual header
    row has two or more non-null cells.
    """
    for i, row in enumerate(rows[:5]):
        if sum(1 for v in row if v is not None) >= 2:
            return i
    return 0


def _read_rows(path: Path, table_name: str) -> tuple[list[str], list[list[Any]]]:
    suffix = path.suffix.lower()
    if suffix in (".xls", ".xlsx"):
        if not _HAS_OPENPYXL:
            raise RuntimeError("openpyxl is required for Excel files: pip install openpyxl")
        wb = _openpyxl.load_workbook(path, read_only=True, data_only=True)  # type: ignore[union-attr]
        sheet_name = next(
            (s for s in wb.sheetnames if _normalize_sheet_name(s) == table_name),
            None,
        )
        if sheet_name is None:
            raise KeyError(f"No sheet matching table {table_name!r} in {path.name} (sheets: {wb.sheetnames})")
        all_rows = list(wb[sheet_name].iter_rows(values_only=True))
        header_idx = _find_header_row(all_rows)
        return [str(h) if h is not None else "" for h in all_rows[header_idx]], [
            list(r) for r in all_rows[header_idx + 1 :]
        ]
    with open(path, newline="") as f:
        rows = list(csv.reader(f))
    return rows[0], rows[1:]


def _load_rows(path: Path, schema: TargetSchema) -> dict[str, list[dict[str, Any]] | None]:
    """Return table_name → row dicts, or None when required columns are absent.

    Matching strategy (in order):
    1. Normalized name match — ``FY2022`` matches schema column ``fy2022``.
    2. Positional fallback — when the LLM semantically renamed a column
       (e.g. "Fiscal Year" → ``line_item``), fall back to aligning schema
       columns to file columns by position, provided the counts match.
    """
    rows_by_table: dict[str, list[dict[str, Any]] | None] = {}
    for table in schema.tables:
        headers, raw_rows = _read_rows(path, table.name)
        # Build normalized-name → index map (skip empty/None header cells).
        norm_header_idx: dict[str, int] = {}
        for i, h in enumerate(headers):
            nk = _normalize_col(h)
            if nk and nk not in norm_header_idx:
                norm_header_idx[nk] = i

        col_names = [c.name for c in table.columns]
        missing = [c for c in col_names if _normalize_col(c) not in norm_header_idx]

        if missing:
            # Positional fallback: align schema column i to file column i.
            # First try matching the total column count (including None-header cells,
            # common in multi-section sheets like dashboards).  If that doesn't fit,
            # try matching only non-empty header positions.
            if len(headers) >= len(col_names):
                col_idx = {c: i for i, c in enumerate(col_names)}
            else:
                non_empty_headers = [h for h in headers if h and h.strip()]
                if len(non_empty_headers) >= len(col_names):
                    col_idx = {c: headers.index(non_empty_headers[i]) for i, c in enumerate(col_names)}
                else:
                    rows_by_table[table.name] = None
                    continue
        else:
            col_idx = {c: norm_header_idx[_normalize_col(c)] for c in col_names}

        rows_by_table[table.name] = [
            {c: (row[col_idx[c]] if col_idx[c] < len(row) else None) for c in col_names} for row in raw_rows
        ]
    return rows_by_table


def _quote(name: str) -> str:
    return '"' + name.replace('"', '""') + '"'


def _sync_ingest_table(
    db_path: Path,
    table_spec: TableSpec,
    rows: list[dict[str, Any]],
) -> dict[str, Any]:
    conn = sqlite3.connect(db_path)
    try:
        col_defs: list[str] = []
        for col in table_spec.columns:
            parts = [_quote(col.name), _SQL_TYPES[col.type]]
            if col.primary_key:
                parts.append("PRIMARY KEY")
            elif not col.nullable:
                parts.append("NOT NULL")
            col_defs.append(" ".join(parts))
        conn.execute(f"CREATE TABLE IF NOT EXISTS {_quote(table_spec.name)} ({', '.join(col_defs)})")
        col_names = [c.name for c in table_spec.columns]
        placeholders = ", ".join("?" for _ in col_names)
        col_names_str = ", ".join(_quote(c) for c in col_names)
        errors: list[str] = []
        inserted = 0
        for row_num, row in enumerate(rows, start=2):
            values = [row.get(c) for c in col_names]
            # Skip section-header rows: all values are None (common in Excel).
            if all(v is None for v in values):
                continue
            try:
                conn.execute(
                    f"INSERT INTO {_quote(table_spec.name)} ({col_names_str}) VALUES ({placeholders})",
                    values,
                )
                inserted += 1
            except sqlite3.Error as exc:
                errors.append(f"row {row_num}: {exc}")
        if inserted == 0 and errors:
            conn.rollback()
            return {"status": "failed", "inserted": 0, "errors": errors}
        conn.commit()
        status = "partial" if errors else "success"
        return {"status": status, "inserted": inserted, "errors": errors}
    finally:
        conn.close()


async def ingest_structured(
    path: Path,
    db_path: Path,
    schema: TargetSchema,
) -> dict[str, Any]:
    """Insert rows from *path* into *db_path* according to *schema*.

    Returns ``{table_name: {status, inserted, errors}}`` for each table.
    Missing columns are reported without aborting other tables.
    """
    rows_by_table = _load_rows(path, schema)
    loop = asyncio.get_running_loop()
    results: dict[str, Any] = {}
    for table_spec in schema.tables:
        rows = rows_by_table.get(table_spec.name)
        if rows is None:
            results[table_spec.name] = {
                "status": "failed",
                "inserted": 0,
                "errors": [f"missing columns for table {table_spec.name!r}"],
            }
            continue
        results[table_spec.name] = await loop.run_in_executor(None, _sync_ingest_table, db_path, table_spec, rows)
    return results
